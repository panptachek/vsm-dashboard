#!/usr/bin/env python3
"""Import VSM project plan spreadsheets into works_db_v2.

Inputs are expected to be converted to XLSX in /tmp/vsm_import_inputs:
  - Сваи на май 2026.xlsx
  - Попикетная ведомость ОХ.xlsx
  - Попикетная ведомость ВАД.xlsx
"""

from __future__ import annotations

import argparse
import re
import subprocess
import uuid
from collections import defaultdict
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


INPUT_DIR = Path("/tmp/vsm_import_inputs")
PILE_FILE = INPUT_DIR / "Сваи на май 2026.xlsx"
MAIN_FILE = INPUT_DIR / "Попикетная ведомость ОХ.xlsx"
TEMP_ROAD_FILE = INPUT_DIR / "Попикетная ведомость ВАД.xlsx"

SOURCE_PILES = "Сваи на май 2026.xlsb"
SOURCE_MAIN = "Попикетная ведомость ОХ.xls"
SOURCE_TEMP_ROADS = "Попикетная ведомость ВАД.xlsx"

TEMP_ROAD_OBJECTS = {
    "АД1": "VPD_001",
    "АД3": "VPD_003",
    "АД5": "VPD_005",
    "АД6": "VPD_006",
    "АД7": "VPD_007",
    "АД9": "VPD_009",
    "АД11": "VPD_011",
    "АД12": "VPD_012",
    "АД13": "VPD_013",
    "АД14": "VPD_014",
    "АД15": "VPD_015",
    "АД2 №6": "VPD_026",
    "АД2 №7": "VPD_027",
    "АД4 №7": "VPD_047",
    "АД4 №8": "VPD_048",
    "АД4 №8.1": "VPD_481",
    "АД4 №9": "VPD_049",
    "АД8 №1": "VPD_081",
    "АД8 №2": "VPD_082",
}

TEMP_ROAD_WORKS = {
    "Насыпь, м3": ("EMBANKMENT_CONSTRUCTION", "Устройство насыпи", "м3"),
    "Замена грунта, м3": ("WEAK_SOIL_REPLACEMENT", "Замена слабого грунта", "м3"),
    "Выемка, м3": ("EARTH_EXCAVATION", "Разработка выемки", "м3"),
    "Устройство кюветов, м3": ("DITCH_CONSTRUCTION", "Устройство кюветов", "м3"),
    "Устройство слоя дорожной одежды из песка, м3": ("PAVEMENT_SANDING", "Устройство слоя ДО из песка", "м3"),
    "Отсыпка присыпных обочин из песка, м3": ("SHOULDER_BACKFILL", "Отсыпка присыпных обочин", "м3"),
}

MAIN_WORKS = {
    "Выемка, м3": ("EXCAVATION_MAIN", "Выемка ОХ", "м3"),
    "Выторфовка, м3": ("PEAT_REMOVAL", "Выторфовка", "м3"),
    "Насыпь, м3": ("EMBANKMENT_CONSTRUCTION", "Устройство насыпи", "м3"),
    "ЗС2, м3": ("ZS2", "ЗС2", "м3"),
    "ЗС1, м3": ("ZS1", "ЗС1", "м3"),
}


@dataclass(frozen=True)
class WorkKey:
    source: str
    object_code: str
    constructive_code: str
    work_type_code: str
    unit: str


@dataclass
class Segment:
    key: WorkKey
    pk_start: float
    pk_end: float
    volume: float
    comment: str


@dataclass
class PileField:
    id: str
    field_code: str
    field_type: str
    pk_start: float
    pk_end: float
    section_code: str | None


def run_psql(sql: str) -> str:
    result = subprocess.run(
        ["docker", "exec", "-i", "works_postgres_v2", "psql", "-U", "works_user", "-d", "works_db_v2", "-At", "-F", "\t", "-c", sql],
        check=True,
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
    )
    return result.stdout


def run_psql_script(sql: str) -> None:
    subprocess.run(
        ["docker", "exec", "-i", "works_postgres_v2", "psql", "-U", "works_user", "-d", "works_db_v2", "-v", "ON_ERROR_STOP=1"],
        input=sql,
        check=True,
        text=True,
    )


def lit(value: Any) -> str:
    if value is None:
        return "NULL"
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (int, float)):
        return str(value)
    return "'" + str(value).replace("'", "''") + "'"


def num(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def is_positive(value: Any) -> bool:
    return num(value) > 0


def normalize_road(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("№", "№")).strip()


def pk_from_parts(pk: Any, plus: Any) -> float | None:
    if pk is None or plus is None:
        return None
    return float(pk) * 100.0 + float(str(plus).replace(",", "."))


def parse_pk_text(text: str) -> list[float]:
    values: list[float] = []
    # Full forms: ПК 2656+36,50
    for pk, plus in re.findall(r"(?:ПК\s*)?(\d{4})\s*\+\s*(\d+(?:[,.]\d+)?)", text, flags=re.I):
        values.append(float(pk) * 100.0 + float(plus.replace(",", ".")))
    return values


def section_code_from_text(text: Any) -> str | None:
    m = re.search(r"УЧАСТОК\s*№\s*(\d+)", str(text or ""), flags=re.I)
    if not m:
        return None
    n = int(m.group(1))
    return f"UCH_{n}" if 1 <= n <= 8 else None


def field_base(text: str) -> str | None:
    m = re.match(r"\s*(\d+(?:_\d+)?)", text)
    return m.group(1) if m else None


def load_pile_fields() -> list[PileField]:
    sql = """
        SELECT pf.id::text, pf.field_code, pf.field_type, pf.pk_start::float8, pf.pk_end::float8, cs.code
        FROM pile_fields pf
        LEFT JOIN construction_section_versions csv
          ON csv.is_current = true
         AND pf.pk_start >= csv.pk_start
         AND pf.pk_end <= csv.pk_end
        LEFT JOIN construction_sections cs ON cs.id = csv.section_id
        WHERE pf.is_demo = false
    """
    fields: list[PileField] = []
    for line in run_psql(sql).splitlines():
        fid, code, typ, start, end, section = (line.split("\t") + [""])[:6]
        fields.append(PileField(fid, code, typ, float(start), float(end), section or None))
    return fields


def norm_field_code(code: str) -> str:
    return re.sub(r"\s+[НТ]$", "", code.strip(), flags=re.I)


def find_pile_field(fields: list[PileField], section_code: str | None, text: str, kind: str | None) -> str | None:
    base = field_base(text)
    if not base:
        return None
    points = parse_pk_text(text)
    kind_lower = (kind or "").lower()
    preferred_type = "test" if ("испыту" in kind_lower or "динами" in kind_lower) else "main"
    candidates = [f for f in fields if norm_field_code(f.field_code).startswith(base)]
    if section_code:
        scoped = [f for f in candidates if f.section_code == section_code]
        if scoped:
            candidates = scoped
    if points:
        point = points[0]
        containing = [f for f in candidates if f.pk_start - 0.01 <= point <= f.pk_end + 0.01]
        exact = [f for f in containing if f.field_type == preferred_type]
        if exact:
            return exact[0].id
        if containing:
            return containing[0].id
    exact_code = [f for f in candidates if norm_field_code(f.field_code) == base and f.field_type == preferred_type]
    if exact_code:
        return exact_code[0].id
    return candidates[0].id if candidates else None


def parse_temp_road_plans() -> tuple[dict[WorkKey, float], list[Segment], set[str]]:
    wb = load_workbook(TEMP_ROAD_FILE, read_only=True, data_only=True)
    ws = wb["by_vsm_pk"]
    headers = [cell.value for cell in ws[1]]
    idx = {name: headers.index(name) for name in headers if name}
    totals: dict[WorkKey, float] = defaultdict(float)
    segments: list[Segment] = []
    skipped_roads: set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        road = normalize_road(row[idx["№ автодороги"]])
        if not road or road == "Итого:":
            continue
        object_code = TEMP_ROAD_OBJECTS.get(road)
        if not object_code:
            skipped_roads.add(road)
            continue
        pk_start = row[idx["pk_start_vsm_m"]]
        pk_end = row[idx["pk_end_vsm_m"]]
        if pk_start is None or pk_end is None:
            skipped_roads.add(road)
            continue
        for col_name, (work_code, _work_name, unit) in TEMP_ROAD_WORKS.items():
            value = num(row[idx[col_name]])
            if value <= 0:
                continue
            key = WorkKey(SOURCE_TEMP_ROADS, object_code, "VPD", work_code, unit)
            totals[key] += value
            comment = (
                f"{road}; ПК ВСЖМ {row[idx['ПК начало']]} - {row[idx['ПК конец']]}; "
                f"ПК АД {row[idx['pk_start_ad']]} - {row[idx['pk_end_ad']]}; status={row[idx['status']]}"
            )
            segments.append(Segment(key, float(pk_start), float(pk_end), value, comment))
    return totals, segments, skipped_roads


def parse_main_track_plans() -> tuple[dict[WorkKey, float], list[Segment]]:
    wb = load_workbook(MAIN_FILE, read_only=True, data_only=True)
    ws = wb["ИТОГ"]
    totals: dict[WorkKey, float] = defaultdict(float)
    segments: list[Segment] = []
    headers = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
    # The real column names are split across rows 2 and 3.
    work_col = {
        "Выемка, м3": 5,
        "Выторфовка, м3": 6,
        "Насыпь, м3": 7,
        "ЗС2, м3": 8,
        "ЗС1, м3": 9,
    }
    current_start: float | None = None
    current_label: str | None = None
    rows = list(ws.iter_rows(min_row=4, values_only=True))
    for i, row in enumerate(rows):
        boundary = pk_from_parts(row[0], row[1])
        if boundary is not None:
            current_start = boundary
            current_label = str(row[2] or "")
            continue
        if current_start is None:
            continue
        distance = num(row[3])
        if distance <= 0:
            continue
        next_boundary = None
        for future in rows[i + 1 :]:
            next_boundary = pk_from_parts(future[0], future[1])
            if next_boundary is not None:
                break
        pk_end = float(next_boundary if next_boundary is not None else current_start + distance)
        for col_name, (work_code, _work_name, unit) in MAIN_WORKS.items():
            value = num(row[work_col[col_name] - 1])
            if value <= 0:
                continue
            key = WorkKey(SOURCE_MAIN, "MAIN_OH_STAGE3", "MAIN", work_code, unit)
            totals[key] += value
            comment = f"ОХ; ПК ВСЖМ {current_label or current_start} - {pk_end:.2f}; distance={distance:g} м"
            segments.append(Segment(key, current_start, pk_end, value, comment))
    return totals, segments


def parse_pile_plans(fields: list[PileField]) -> tuple[dict[tuple[str, str | None, date], dict[str, Any]], list[str]]:
    wb = load_workbook(PILE_FILE, read_only=True, data_only=True)
    ws = wb["Отчет"]
    day_cols = []
    for col in range(1, ws.max_column + 1):
        value = ws.cell(3, col).value
        if isinstance(value, (int, float)) and 1 <= int(value) <= 31:
            day_cols.append((col, int(value)))

    plans: dict[tuple[str, str | None, date], dict[str, Any]] = defaultdict(
        lambda: {"main": 0, "test": 0, "dynamic": 0, "comments": set()}
    )
    warnings: list[str] = []
    current_section: str | None = None
    active_field_text: str | None = None
    active_kind: str | None = None
    active_field_id: str | None = None

    for row_num in range(4, ws.max_row + 1):
        row = [ws.cell(row_num, col).value for col in range(1, 8)]
        first = str(row[0] or "")
        field_text = str(row[1] or "")
        if "УЧАСТОК" in field_text.upper():
            current_section = section_code_from_text(field_text)
            active_field_text = None
            active_kind = None
            active_field_id = None
            continue
        if "ИТОГО" in first.upper():
            active_field_text = None
            active_kind = None
            active_field_id = None
            continue
        is_detail_row = bool(re.match(r"^\s*\d+\.\d+", first)) and bool(field_text)
        if is_detail_row:
            active_field_text = field_text
            active_kind = str(row[2] or "")
            active_field_id = find_pile_field(fields, current_section, active_field_text, active_kind)
            if not active_field_id:
                warnings.append(f"Не найдено pile_field для строки {row_num}: {active_field_text}")

        marker = str(row[5] or "").strip().lower()
        if marker != "план" or not current_section or not active_field_text:
            continue
        kind = active_kind or str(row[2] or "")
        machine = str(row[3] or "").strip()
        for col, day in day_cols:
            raw = ws.cell(row_num, col).value
            main = test = dyn = 0
            if isinstance(raw, (int, float)) and raw > 0:
                if "испыту" in kind.lower():
                    test = int(round(raw))
                elif "динами" in kind.lower():
                    dyn = int(round(raw))
                else:
                    main = int(round(raw))
            elif isinstance(raw, str) and raw.strip().upper().startswith("Д"):
                m = re.search(r"(\d+)", raw)
                dyn = int(m.group(1)) if m else 1
            if main == test == dyn == 0:
                continue
            d = date(2026, 5, day)
            key = (current_section, active_field_id, d)
            plans[key]["main"] += main
            plans[key]["test"] += test
            plans[key]["dynamic"] += dyn
            plans[key]["comments"].add(f"{active_field_text}; {kind or 'вид работ н/д'}; {machine or 'машина н/д'}")
    return plans, warnings


def build_sql(
    project_totals: dict[WorkKey, float],
    segments: list[Segment],
    pile_plans: dict[tuple[str, str | None, date], dict[str, Any]],
) -> str:
    pwi_ids = {key: str(uuid.uuid4()) for key in project_totals}
    lines: list[str] = [
        "BEGIN;",
        "ALTER TABLE pile_plan_periods ADD COLUMN IF NOT EXISTS planned_dynamic_tests integer NOT NULL DEFAULT 0;",
        "DO $$ BEGIN",
        "  IF NOT EXISTS (SELECT 1 FROM pg_constraint WHERE conname = 'pile_plan_periods_dynamic_nonnegative') THEN",
        "    ALTER TABLE pile_plan_periods ADD CONSTRAINT pile_plan_periods_dynamic_nonnegative CHECK (planned_dynamic_tests >= 0) NOT VALID;",
        "  END IF;",
        "END $$;",
        "ALTER TABLE pile_plan_periods VALIDATE CONSTRAINT pile_plan_periods_dynamic_nonnegative;",
        "CREATE TABLE IF NOT EXISTS project_work_item_segments (",
        "  id UUID PRIMARY KEY DEFAULT gen_random_uuid(),",
        "  project_work_item_id UUID NOT NULL REFERENCES project_work_items(id) ON DELETE CASCADE,",
        "  pk_start NUMERIC NOT NULL,",
        "  pk_end NUMERIC NOT NULL,",
        "  pk_raw_text TEXT,",
        "  volume_segment NUMERIC,",
        "  comment TEXT,",
        "  created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),",
        "  UNIQUE (project_work_item_id, pk_start, pk_end)",
        ");",
        "CREATE INDEX IF NOT EXISTS idx_project_work_item_segments_item ON project_work_item_segments(project_work_item_id);",
        "CREATE INDEX IF NOT EXISTS idx_project_work_item_segments_pk ON project_work_item_segments(pk_start, pk_end);",
        "ALTER TABLE project_work_item_segments ADD COLUMN IF NOT EXISTS pk_raw_text TEXT;",
        "ALTER TABLE project_work_item_segments ADD COLUMN IF NOT EXISTS volume_segment NUMERIC;",
        "DO $$ BEGIN",
        "  IF EXISTS (SELECT 1 FROM information_schema.columns WHERE table_name = 'project_work_item_segments' AND column_name = 'project_volume_segment') THEN",
        "    UPDATE project_work_item_segments SET volume_segment = project_volume_segment WHERE volume_segment IS NULL;",
        "  END IF;",
        "END $$;",
        "ALTER TABLE project_work_item_segments DROP COLUMN IF EXISTS project_volume_segment;",
        "INSERT INTO work_types (id, code, name, default_unit, work_group, is_active, show_in_timeline)",
        "VALUES",
        "  (gen_random_uuid(), 'ZS1', 'ЗС1', 'м3', 'earthwork', true, true),",
        "  (gen_random_uuid(), 'ZS2', 'ЗС2', 'м3', 'earthwork', true, true)",
        "ON CONFLICT (code) DO UPDATE SET name = EXCLUDED.name, default_unit = EXCLUDED.default_unit, work_group = EXCLUDED.work_group, is_active = true, show_in_timeline = true;",
        "UPDATE work_types SET default_unit = 'м3' WHERE code = 'EXCAVATION_MAIN';",
        "INSERT INTO objects (id, object_code, name, object_type_id, constructive_id, is_active, comment)",
        "VALUES (gen_random_uuid(), 'MAIN_OH_STAGE3', 'Основной ход ВСЖМ-1, этап 3',",
        "        (SELECT id FROM object_types WHERE code = 'MAIN_TRACK'),",
        "        (SELECT id FROM constructives WHERE code = 'MAIN'),",
        "        true, 'auto-created for main track project volume import')",
        "ON CONFLICT (object_code) DO UPDATE SET name = EXCLUDED.name, object_type_id = EXCLUDED.object_type_id, constructive_id = EXCLUDED.constructive_id, is_active = true;",
        f"DELETE FROM pile_plan_periods WHERE source_reference = {lit(SOURCE_PILES)};",
        f"DELETE FROM project_work_item_segments WHERE project_work_item_id IN (SELECT id FROM project_work_items WHERE source_reference IN ({lit(SOURCE_MAIN)}, {lit(SOURCE_TEMP_ROADS)}));",
        f"DELETE FROM project_work_items WHERE source_reference IN ({lit(SOURCE_MAIN)}, {lit(SOURCE_TEMP_ROADS)});",
    ]

    for key, volume in sorted(project_totals.items(), key=lambda item: (item[0].source, item[0].object_code, item[0].work_type_code)):
        lines.append(
            "INSERT INTO project_work_items (id, object_id, constructive_id, work_type_id, project_volume, unit, source_reference, comment) "
            f"VALUES ({lit(pwi_ids[key])}::uuid, "
            f"(SELECT id FROM objects WHERE object_code = {lit(key.object_code)}), "
            f"(SELECT id FROM constructives WHERE code = {lit(key.constructive_code)}), "
            f"(SELECT id FROM work_types WHERE code = {lit(key.work_type_code)}), "
            f"{volume:.6f}, {lit(key.unit)}, {lit(key.source)}, "
            f"{lit('imported project total from source spreadsheet')});"
        )

    for seg in segments:
        lines.append(
            "INSERT INTO project_work_item_segments (id, project_work_item_id, pk_start, pk_end, volume_segment, comment) "
            f"VALUES ({lit(str(uuid.uuid4()))}::uuid, {lit(pwi_ids[seg.key])}::uuid, "
            f"{seg.pk_start:.6f}, {seg.pk_end:.6f}, {seg.volume:.6f}, {lit(seg.comment)});"
        )

    for (section_code, field_id, d), data in sorted(pile_plans.items(), key=lambda item: (item[0][0], item[0][2], item[0][1] or "")):
        comments = "; ".join(sorted(data["comments"]))[:1800]
        lines.append(
            "INSERT INTO pile_plan_periods (id, section_id, pile_field_id, period_start, period_end, plan_type, "
            "planned_main_piles, planned_test_piles, planned_dynamic_tests, source_reference, comment) "
            f"VALUES ({lit(str(uuid.uuid4()))}::uuid, "
            f"(SELECT id FROM construction_sections WHERE code = {lit(section_code)}), "
            f"{lit(field_id + '::uuid') if False else ('NULL' if field_id is None else lit(field_id) + '::uuid')}, "
            f"{lit(d.isoformat())}::date, {lit(d.isoformat())}::date, 'daily', "
            f"{int(data['main'])}, {int(data['test'])}, {int(data['dynamic'])}, {lit(SOURCE_PILES)}, {lit(comments)});"
        )
    lines.append("COMMIT;")
    return "\n".join(lines)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--apply", action="store_true", help="Apply import to the database")
    parser.add_argument("--sql-out", default="/tmp/vsm_project_plan_import.sql")
    args = parser.parse_args()

    for path in [PILE_FILE, MAIN_FILE, TEMP_ROAD_FILE]:
        if not path.exists():
            raise SystemExit(f"Missing input: {path}")

    fields = load_pile_fields()
    temp_totals, temp_segments, skipped_roads = parse_temp_road_plans()
    main_totals, main_segments = parse_main_track_plans()
    pile_plans, warnings = parse_pile_plans(fields)
    totals = defaultdict(float)
    totals.update(temp_totals)
    for key, value in main_totals.items():
        totals[key] += value
    segments = temp_segments + main_segments
    sql = build_sql(dict(totals), segments, pile_plans)
    Path(args.sql_out).write_text(sql, encoding="utf-8")

    work_names = {}
    for source_map in (TEMP_ROAD_WORKS, MAIN_WORKS):
        for _col, (code, name, unit) in source_map.items():
            work_names[code] = f"{name}, {unit}"
    summary = {
        "project_work_items": len(totals),
        "project_work_item_segments": len(segments),
        "temp_road_segments": len(temp_segments),
        "main_track_segments": len(main_segments),
        "pile_plan_rows": len(pile_plans),
        "pile_plan_main": sum(v["main"] for v in pile_plans.values()),
        "pile_plan_test": sum(v["test"] for v in pile_plans.values()),
        "pile_plan_dynamic": sum(v["dynamic"] for v in pile_plans.values()),
        "skipped_temp_roads": sorted(skipped_roads),
        "pile_mapping_warnings": warnings[:30],
        "work_names": work_names,
        "sql_out": args.sql_out,
    }
    print(summary)
    if args.apply:
        run_psql_script(sql)
        print({"applied": True, "sql_out": args.sql_out})


if __name__ == "__main__":
    main()
