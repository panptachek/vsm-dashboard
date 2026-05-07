"""Convert the VSM daily XLSX template into the deterministic text format.

The dashboard parser intentionally stays text-first. This adapter is a thin
compatibility layer for the approved workbook template:

    Отчет шаблон ВСЖМ (для текстовой части).xlsx

It performs no DB writes and does not interpret values beyond reshaping workbook
rows into the existing ``===Секция===`` report text contract.
"""
from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import load_workbook


TEMPLATE_MARKER = "ДЕЖУРНЫЙ ОТЧЁТ ПО УЧАСТКУ"


def _clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).replace("\xa0", " ").strip()


def _has_value(*values: Any) -> bool:
    return any(_clean(value) and _clean(value).lower() not in {"н/д", "нд", "-"} for value in values)


def _sheet_score(ws) -> int:
    score = 0
    if TEMPLATE_MARKER.lower() in _clean(ws["A1"].value).lower():
        score += 100
    if _has_value(ws["A5"].value, ws["B5"].value, ws["C5"].value):
        score += 50
    data_ranges = (
        ("A", "L", 32, 76),
        ("A", "N", 80, 104),
        ("A", "M", 108, 122),
        ("A", "F", 126, 180),
        ("A", "E", 195, 204),
        ("A", "G", 208, 217),
    )
    for start_col, end_col, start_row, end_row in data_ranges:
        start_idx = ws[start_col + str(start_row)].column
        end_idx = ws[end_col + str(start_row)].column
        for row_idx in range(start_row, end_row + 1):
            if _has_value(*(ws.cell(row_idx, col_idx).value for col_idx in range(start_idx, end_idx + 1))):
                score += 1
    return score


def _content_score(ws) -> int:
    score = _sheet_score(ws)
    if TEMPLATE_MARKER.lower() in _clean(ws["A1"].value).lower():
        score -= 100
    return score


def _select_report_sheet(workbook):
    candidates = [
        ws
        for ws in workbook.worksheets
        if "справоч" not in ws.title.lower()
        and TEMPLATE_MARKER.lower() in _clean(ws["A1"].value).lower()
    ]
    if not candidates:
        candidates = [ws for ws in workbook.worksheets if "справоч" not in ws.title.lower()]
    if not candidates:
        return workbook.active
    filled = [ws for ws in candidates if _content_score(ws) > 0]
    active = workbook.active
    if active in filled and not active.title.lower().startswith("пример"):
        return active
    for ws in filled:
        if ws.title.strip().lower() == "шаблон":
            return ws
    non_example = [ws for ws in filled if not ws.title.lower().startswith("пример")]
    if non_example:
        return max(non_example, key=_sheet_score)
    if filled:
        return max(filled, key=_sheet_score)
    return max(candidates, key=_sheet_score)


def _vehicle_line(kind: Any, model: Any, plate: Any, owner: Any) -> str:
    kind_text = _clean(kind) or "Техника"
    model_text = _clean(model)
    left = " ".join(part for part in (kind_text, model_text) if part)
    owner_text = _clean(owner) or "Н/Д"
    return f"{left} (б/н ; г/н {_clean(plate)}); {owner_text}"


def _pk(value: Any) -> str:
    text = _clean(value)
    if not text or text.lower() in {"н/д", "нд", "-"}:
        return ""
    return text if text.lower().startswith("пк") else f"ПК{text}"


def _pk_range(start: Any, end: Any) -> str:
    left = _pk(start)
    right = _pk(end)
    if left and right:
        return f"{left}-{right}"
    return left or right or "Н/Д"


def _volume(value: Any) -> str:
    text = _clean(value)
    return text if text else "Н/Д"


def _append_comment(lines: list[str], comment: Any) -> None:
    text = _clean(comment)
    if text:
        lines.append(f"%% {text}")


def _append_transport(lines: list[str], ws) -> None:
    lines.extend(["", "===Перевозка==="])
    current_key: tuple[str, str, str, str, str, str] | None = None
    for row_idx in range(32, 77):
        driver = _clean(ws[f"A{row_idx}"].value)
        kind = _clean(ws[f"B{row_idx}"].value)
        model = _clean(ws[f"C{row_idx}"].value)
        plate = _clean(ws[f"D{row_idx}"].value)
        owner = _clean(ws[f"E{row_idx}"].value)
        comment = _clean(ws[f"F{row_idx}"].value)
        material = _clean(ws[f"G{row_idx}"].value)
        source = _clean(ws[f"H{row_idx}"].value)
        destination = _clean(ws[f"I{row_idx}"].value)
        volume = _clean(ws[f"J{row_idx}"].value)
        trips = _clean(ws[f"K{row_idx}"].value)
        unit = _clean(ws[f"L{row_idx}"].value) or "м3"
        if not _has_value(driver, kind, model, plate, owner, material, source, destination, volume, trips):
            continue

        key = (driver or "Н/Д", kind, model, plate, owner, comment)
        if key != current_key:
            lines.extend(["", f"ФИО: {driver or 'Н/Д'}"])
            lines.append(_vehicle_line(kind, model, plate, owner))
            _append_comment(lines, comment)
            current_key = key
        if material:
            lines.append(f"/{material}/")
        if source or destination:
            lines.append(f"{source or 'Н/Д'} → {destination or 'Н/Д'}")
        if volume or trips:
            lines.append(f"{_volume(volume)} {unit} / {trips or 0} рейс")


def _append_work_row(
    lines: list[str],
    *,
    constructive: Any = None,
    work_name: Any,
    rail_start: Any,
    rail_end: Any,
    ad_start: Any,
    ad_end: Any,
    operator: Any,
    equipment_kind: Any,
    model: Any,
    plate: Any,
    owner: Any,
    volume: Any,
    unit: Any,
    comment: Any,
) -> None:
    if not _has_value(work_name, rail_start, rail_end, ad_start, ad_end, operator, equipment_kind, model, plate, owner, volume, unit, comment):
        return
    constructive_text = _clean(constructive)
    if constructive_text:
        lines.append(f"-{constructive_text}-")
    lines.append(f"/{_clean(work_name) or 'Н/Д'}/")
    lines.append(f"ПК ВСЖМ: {_pk_range(rail_start, rail_end)}")
    lines.append(f"ПК АД: {_pk_range(ad_start, ad_end)}")
    operator_text = _clean(operator)
    if operator_text:
        lines.append(f"ФИО: {operator_text}")
    if _has_value(equipment_kind, model, plate, owner):
        lines.append(_vehicle_line(equipment_kind, model, plate, owner))
    if _has_value(volume, unit):
        lines.append(f"V = {_volume(volume)} {_clean(unit) or 'м3'}")
    _append_comment(lines, comment)
    lines.append("")


def _append_main_works(lines: list[str], ws) -> None:
    lines.extend(["", "===Основные работы==="])
    for row_idx in range(80, 105):
        _append_work_row(
            lines,
            constructive=ws[f"A{row_idx}"].value,
            work_name=ws[f"B{row_idx}"].value,
            rail_start=ws[f"C{row_idx}"].value,
            rail_end=ws[f"D{row_idx}"].value,
            ad_start=ws[f"E{row_idx}"].value,
            ad_end=ws[f"F{row_idx}"].value,
            operator=ws[f"G{row_idx}"].value,
            equipment_kind=ws[f"H{row_idx}"].value,
            model=ws[f"I{row_idx}"].value,
            plate=ws[f"J{row_idx}"].value,
            owner=ws[f"K{row_idx}"].value,
            volume=ws[f"L{row_idx}"].value,
            unit=ws[f"M{row_idx}"].value,
            comment=ws[f"N{row_idx}"].value,
        )


def _append_aux_works(lines: list[str], ws) -> None:
    lines.extend(["", "===Сопутствующие работы==="])
    for row_idx in range(108, 123):
        _append_work_row(
            lines,
            work_name=ws[f"A{row_idx}"].value,
            rail_start=ws[f"B{row_idx}"].value,
            rail_end=ws[f"C{row_idx}"].value,
            ad_start=ws[f"D{row_idx}"].value,
            ad_end=ws[f"E{row_idx}"].value,
            operator=ws[f"F{row_idx}"].value,
            equipment_kind=ws[f"G{row_idx}"].value,
            model=ws[f"H{row_idx}"].value,
            plate=ws[f"I{row_idx}"].value,
            owner=ws[f"J{row_idx}"].value,
            volume=ws[f"K{row_idx}"].value,
            unit=ws[f"L{row_idx}"].value,
            comment=ws[f"M{row_idx}"].value,
        )


def _append_park(lines: list[str], ws) -> None:
    lines.extend(["", "===Парк техники==="])
    for row_idx in range(126, 181):
        kind = ws[f"A{row_idx}"].value
        model = ws[f"B{row_idx}"].value
        plate = ws[f"C{row_idx}"].value
        owner = ws[f"D{row_idx}"].value
        status = _clean(ws[f"E{row_idx}"].value)
        comment = _clean(ws[f"F{row_idx}"].value)
        if not _has_value(kind, model, plate, owner, status, comment):
            continue
        tail = "; ".join(part for part in (_clean(owner) or "Н/Д", status, comment) if part)
        left = " ".join(part for part in (_clean(kind) or "Техника", _clean(model)) if part)
        lines.append(f"{left} (б/н ; г/н {_clean(plate)}); {tail}")


def _append_problems(lines: list[str], ws) -> None:
    lines.extend(["", "===Проблемные вопросы==="])
    for row_idx in range(183, 188):
        text = _clean(ws[f"A{row_idx}"].value)
        if text:
            lines.append(text)


def _append_personnel(lines: list[str], ws) -> None:
    lines.extend(["", "===Персонал==="])
    mapping = (
        ("ДР", ws["A191"].value),
        ("ИТР", ws["B191"].value),
        ("Водители", ws["C191"].value),
        ("Механизаторы", ws["D191"].value),
    )
    for label, value in mapping:
        text = _clean(value)
        if text and text.lower() not in {"н/д", "нд", "-"}:
            lines.append(f"{label}: {text}")
    _append_comment(lines, ws["E191"].value)


def _append_stockpiles(lines: list[str], ws) -> None:
    lines.extend(["", "===Накопители==="])
    for row_idx in range(195, 205):
        name = _clean(ws[f"A{row_idx}"].value)
        pk = _pk(ws[f"B{row_idx}"].value)
        volume = _clean(ws[f"C{row_idx}"].value)
        unit = _clean(ws[f"D{row_idx}"].value) or "м3"
        comment = _clean(ws[f"E{row_idx}"].value)
        if not _has_value(name, pk, volume, comment):
            continue
        line = f"{name or 'Накопитель'} - {pk or 'Н/Д'} - {_volume(volume)} {unit}"
        if comment:
            line = f"{line} %% {comment}"
        lines.append(line)


def _append_pile_driving(lines: list[str], ws) -> None:
    lines.extend(["", "===Забивка свай==="])
    for row_idx in range(208, 218):
        field_code = _clean(ws[f"A{row_idx}"].value)
        pk = _pk(ws[f"B{row_idx}"].value)
        pile_kind_raw = _clean(ws[f"C{row_idx}"].value)
        count = _clean(ws[f"D{row_idx}"].value)
        pile_type = _clean(ws[f"E{row_idx}"].value)
        composite = _clean(ws[f"F{row_idx}"].value)
        comment = _clean(ws[f"G{row_idx}"].value)
        if not _has_value(field_code, pk, count, pile_type, composite, comment):
            continue
        pile_kind = pile_kind_raw or "основные"
        parts = [field_code or "поле н/д", pk or "ПК н/д", pile_kind]
        if count:
            parts.append(f"{count} шт")
        if pile_type:
            parts.append(pile_type)
        if composite.lower() in {"да", "yes", "true", "1", "готова", "составная готова"}:
            parts.append("составная готова")
        elif composite:
            parts.append(composite)
        if comment:
            parts.append(comment)
        lines.append(" - ".join(parts))


def workbook_to_report_text(workbook) -> str:
    ws = _select_report_sheet(workbook)
    lines = [
        f"Дата - {_clean(ws['A5'].value)}",
        f"Смена - {_clean(ws['B5'].value)}",
        f"Участок - {_clean(ws['C5'].value)}",
        f"Направление - {_clean(ws['D5'].value)}",
        f"Конструктивы - {_clean(ws['E5'].value)}",
        f"%% XLSX source sheet: {ws.title}",
    ]
    delivery_info = _clean(ws["B6"].value)
    if delivery_info:
        lines.extend(["|", f"Информация по завозу: {delivery_info}"])
    _append_transport(lines, ws)
    _append_main_works(lines, ws)
    _append_aux_works(lines, ws)
    _append_park(lines, ws)
    _append_problems(lines, ws)
    _append_personnel(lines, ws)
    _append_stockpiles(lines, ws)
    _append_pile_driving(lines, ws)
    return "\n".join(lines).strip() + "\n"


def xlsx_bytes_to_report_text(blob: bytes) -> str:
    workbook = load_workbook(BytesIO(blob), data_only=True, read_only=False)
    return workbook_to_report_text(workbook)
