#!/usr/bin/env python3
from __future__ import annotations

from openpyxl import Workbook

from report_text_parser import parse_report_text
from xlsx_report_adapter import workbook_to_report_text


def build_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Шаблон"
    ref = wb.create_sheet("Справочники")
    example = wb.create_sheet("Пример уч.7 день")

    for sheet in (ws, example):
        sheet["A1"] = "ДЕЖУРНЫЙ ОТЧЁТ ПО УЧАСТКУ"
    ref["A1"] = "Справочник"

    # A populated Шаблон must be selected before the demo sheet.
    ws["A5"] = "06.05.2026"
    ws["B5"] = "день"
    ws["C5"] = "участок №1"
    ws["D5"] = "ЗП"
    ws["E5"] = "АД 2.2"
    ws["B6"] = "Карьер Васильки - песок - 666 м3"

    ws["A32"] = "Тестов"
    ws["B32"] = "Самосвал"
    ws["C32"] = "FAW"
    ws["D32"] = "999"
    ws["E32"] = "ЖДС"
    ws["G32"] = "песок"
    ws["H32"] = "карьер Великий"
    ws["I32"] = "АД 2.2"
    ws["J32"] = 10
    ws["K32"] = 1
    ws["L32"] = "м3"

    ws["A80"] = "АД 2.2"
    ws["B80"] = "Устройство насыпи"
    ws["C80"] = "ПК100+00"
    ws["D80"] = "ПК101+00"
    ws["G80"] = "Машинист"
    ws["H80"] = "Бульдозер"
    ws["I80"] = "SEM"
    ws["J80"] = "123"
    ws["K80"] = "ЖДС"
    ws["L80"] = 20
    ws["M80"] = "м3"

    ws["A191"] = 3
    ws["B191"] = 4
    ws["C191"] = 5
    ws["D191"] = 6

    ws["A195"] = "Накопитель песка АД13"
    ws["B195"] = "ПК2715+00"
    ws["C195"] = 1000
    ws["D195"] = "м3"

    ws["A208"] = "5_1 Н"
    ws["B208"] = "ПК2655+73.50"
    ws["C208"] = "основные"
    ws["D208"] = 4
    ws["F208"] = "да"

    example["A5"] = "16.04.2026"
    return wb


def main() -> int:
    text = workbook_to_report_text(build_workbook())
    assert "%% XLSX source sheet: Шаблон" in text
    parsed = parse_report_text("xlsx-test.xlsx", text)
    assert parsed["header"]["report_date"] == "2026-05-06"
    assert parsed["header"]["shift"] == "day"
    assert parsed["header"]["section_code"] == "UCH_1"
    assert "Карьер Васильки" in parsed["human_summary"]["delivery_info"]
    assert len(parsed["transport"]) == 1
    assert parsed["transport"][0]["trips"][0]["material_code"] == "SAND"
    assert len(parsed["main_works"]) == 1
    assert parsed["main_works"][0]["work_type_code"] == "EMBANKMENT_CONSTRUCTION"
    assert [row["category"] for row in parsed["personnel"]] == ["ДР", "ИТР", "Водители", "Механизаторы"]
    assert len(parsed["stockpiles"]) == 1
    assert parsed["stockpiles"][0]["rounded_pk"] == 2715
    assert len(parsed["piles"]) == 1
    assert parsed["piles"][0]["field_code"] == "5_1 Н"
    assert parsed["piles"][0]["is_composite_complete"] is True
    print("ok")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
